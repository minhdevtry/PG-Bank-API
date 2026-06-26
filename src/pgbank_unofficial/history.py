"""Transaction History module — query, search, categorize, export.

Provides high-level operations on top of :meth:`PGBankClient.get_transaction_history`:

- :class:`HistoryQuery` — filter criteria with validation
- :class:`TransactionHistory` — query, search, categorize for a single account
- :func:`query_all_accounts` — parallel fetch across all managed accounts
- :func:`export_to_csv`, :func:`export_to_json`, :func:`export_to_excel` — output

Example:
    >>> from pgbank_unofficial import (
    ...     PGBankClient, TransactionHistory, HistoryQuery,
    ...     export_to_csv, export_to_excel,
    ... )
    >>> from decimal import Decimal
    >>> from datetime import date
    >>> with PGBankClient(username="...", password="...", browser_id="...") as c:
    ...     history = TransactionHistory(c)
    ...     q = HistoryQuery(
    ...         from_date=date(2024, 1, 1),
    ...         to_date=date(2024, 12, 31),
    ...         min_amount=Decimal("100000"),
    ...     )
    ...     transactions = history.query(q)
    ...     rules = {"GRAB": "Di chuyển", "VNG": "Công nghệ"}
    ...     categorized = history.categorize(transactions, rules)
    ...     export_to_excel(categorized, "2024_report.xlsx")
"""

from __future__ import annotations

import csv
import json
import logging
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from decimal import Decimal
from pathlib import Path
from typing import TYPE_CHECKING, Any, Optional, Union

from pgbank_unofficial.models import Transaction, TransactionDirection

if TYPE_CHECKING:
    from pgbank_unofficial.client import PGBankClient
    from pgbank_unofficial.manager import PGBankManager

logger = logging.getLogger(__name__)


# ──────────────────────────────────────────────────────────────────────────────
# Query dataclass
# ──────────────────────────────────────────────────────────────────────────────


@dataclass
class HistoryQuery:
    """Filter criteria for transaction history queries.

    All fields are optional. Unset fields are not applied as filters.
    Multiple filters are AND-combined.

    Args:
        account_number: Specific account to query (None = primary account).
        from_date: Start date (inclusive). None = 30 days ago.
        to_date: End date (inclusive). None = today.
        min_amount: Minimum transaction amount (inclusive).
        max_amount: Maximum transaction amount (inclusive).
        direction: DEBIT or CREDIT filter.
        counterparty: Fuzzy match against counterparty_name or counterparty_account.
        description_search: Substring match in description field.
        category: Exact category match (only meaningful for CategorizedTransaction).

    Raises:
        ValueError: If min_amount > max_amount, or from_date > to_date.
    """

    account_number: Optional[str] = None
    from_date: Optional[Union[date, datetime]] = None
    to_date: Optional[Union[date, datetime]] = None
    min_amount: Optional[Decimal] = None
    max_amount: Optional[Decimal] = None
    direction: Optional[TransactionDirection] = None
    counterparty: Optional[str] = None
    description_search: Optional[str] = None
    category: Optional[str] = None

    def __post_init__(self) -> None:
        if self.min_amount is not None and self.max_amount is not None:
            if self.min_amount > self.max_amount:
                raise ValueError(
                    f"min_amount ({self.min_amount}) cannot exceed max_amount ({self.max_amount})"
                )
        if self.from_date is not None and self.to_date is not None:
            if self.from_date > self.to_date:
                raise ValueError(
                    f"from_date ({self.from_date}) cannot be after to_date ({self.to_date})"
                )


# ──────────────────────────────────────────────────────────────────────────────
# CategorizedTransaction — Transaction with category
# ──────────────────────────────────────────────────────────────────────────────


@dataclass
class CategorizedTransaction(Transaction):
    """A Transaction with an assigned category from auto-categorization rules.

    Attributes:
        category: Category name (default "Uncategorized" if no rule matched).
    """

    category: str = "Uncategorized"

    def to_dict(self) -> dict[str, Any]:
        d = super().to_dict()
        d["category"] = self.category
        return d


# ──────────────────────────────────────────────────────────────────────────────
# TransactionHistory — single account
# ──────────────────────────────────────────────────────────────────────────────


class TransactionHistory:
    """High-level transaction history API for a single PGBank account.

    Args:
        client: A logged-in :class:`PGBankClient` instance.

    Example:
        >>> history = TransactionHistory(client)
        >>> q = HistoryQuery(from_date=date(2024, 1, 1), min_amount=Decimal("100000"))
        >>> recent = history.query(q)
    """

    def __init__(self, client: "PGBankClient") -> None:
        self._client = client

    def query(self, q: HistoryQuery) -> list[Transaction]:
        """Fetch and filter transactions.

        Args:
            q: :class:`HistoryQuery` with filters. If ``from_date``/``to_date``
                are unset, defaults to the last 30 days.

        Returns:
            List of :class:`Transaction` objects matching all filters,
            ordered as returned by PGBank (typically timestamp descending).
        """
        account_number = self._resolve_account_number(q.account_number)
        from_date = q.from_date or (datetime.now() - timedelta(days=30))
        to_date = q.to_date or datetime.now()

        transactions = self._client.get_transaction_history(account_number, from_date, to_date)
        return self._apply_filters(transactions, q)

    def search(self, text: str, *, account_number: Optional[str] = None) -> list[Transaction]:
        """Case-insensitive substring search across description, counterparty_name, counterparty_account.

        Args:
            text: Search string (case-insensitive).
            account_number: If provided, restrict search to one account.
                Defaults to the client's primary account.

        Returns:
            List of matching transactions, ordered by timestamp descending.
        """
        if not text:
            return []

        account_number = self._resolve_account_number(account_number)
        # Search covers last 90 days by default
        to_date = datetime.now()
        from_date = to_date - timedelta(days=90)
        transactions = self._client.get_transaction_history(account_number, from_date, to_date)

        needle = text.lower()
        results: list[Transaction] = []
        for tx in transactions:
            if (
                needle in tx.description.lower()
                or needle in tx.counterparty_name.lower()
                or needle in tx.counterparty_account.lower()
            ):
                results.append(tx)
        # Sort by timestamp descending
        results.sort(key=lambda t: t.timestamp, reverse=True)
        return results

    def categorize(
        self, transactions: list[Transaction], rules: dict[str, str]
    ) -> list[CategorizedTransaction]:
        """Apply category rules to a list of transactions.

        Each rule is ``{keyword: category_name}``. A transaction is assigned
        the category of the first matching rule (case-insensitive substring
        match against description, counterparty_name, counterparty_account).
        Unmatched transactions get category "Uncategorized".

        Args:
            transactions: List of :class:`Transaction` objects.
            rules: Dict of ``{keyword: category_name}``.

        Returns:
            List of :class:`CategorizedTransaction` with category field populated.

        Example:
            >>> rules = {"GRAB": "Di chuyển", "VNG": "Công nghệ"}
            >>> categorized = history.categorize(transactions, rules)
        """
        # Lowercase rules for case-insensitive matching
        lowered = {k.lower(): v for k, v in rules.items()}
        result: list[CategorizedTransaction] = []
        for tx in transactions:
            category = "Uncategorized"
            haystack = " ".join(
                [tx.description, tx.counterparty_name, tx.counterparty_account]
            ).lower()
            for keyword, cat in lowered.items():
                if keyword in haystack:
                    category = cat
                    break
            result.append(
                CategorizedTransaction(
                    id=tx.id,
                    account_number=tx.account_number,
                    type=tx.type,
                    amount=tx.amount,
                    currency=tx.currency,
                    counterparty_name=tx.counterparty_name,
                    counterparty_account=tx.counterparty_account,
                    counterparty_bank=tx.counterparty_bank,
                    description=tx.description,
                    timestamp=tx.timestamp,
                    raw=tx.raw,
                    category=category,
                )
            )
        return result

    # ── Internal helpers ──────────────────────────────────────────────────────

    def _resolve_account_number(self, account_number: Optional[str]) -> str:
        """Get account number from query, or fall back to client's primary account."""
        if account_number:
            return account_number
        accounts = self._client.get_accounts()
        if not accounts:
            raise ValueError("No accounts available on client; specify account_number explicitly")
        return accounts[0].account_number

    @staticmethod
    def _apply_filters(transactions: list[Transaction], q: HistoryQuery) -> list[Transaction]:
        """Apply post-fetch filters in-memory."""
        result: list[Transaction] = []
        for tx in transactions:
            if q.min_amount is not None and tx.amount < q.min_amount:
                continue
            if q.max_amount is not None and tx.amount > q.max_amount:
                continue
            if q.direction is not None and tx.type != q.direction:
                continue
            if q.counterparty is not None:
                needle = q.counterparty.lower()
                if (
                    needle not in tx.counterparty_name.lower()
                    and needle not in tx.counterparty_account.lower()
                ):
                    continue
            if q.description_search is not None:
                if q.description_search.lower() not in tx.description.lower():
                    continue
            if q.category is not None:
                # Only CategorizedTransaction has a category field
                cat = getattr(tx, "category", None)
                if cat != q.category:
                    continue
            result.append(tx)
        return result


# ──────────────────────────────────────────────────────────────────────────────
# Multi-account query
# ──────────────────────────────────────────────────────────────────────────────


def query_all_accounts(
    manager: "PGBankManager",
    q: HistoryQuery,
    *,
    max_workers: int = 4,
) -> dict[str, list[Transaction]]:
    """Fetch transactions from all accounts in :class:`PGBankManager` in parallel.

    The ``account_number`` in ``q`` is ignored (each account is queried independently).

    Args:
        manager: :class:`PGBankManager` instance with accounts registered.
        q: :class:`HistoryQuery` filters.
        max_workers: Max concurrent client queries.

    Returns:
        Dict mapping ``nickname -> list[Transaction]``.
    """
    nicknames = manager.list_accounts()
    if not nicknames:
        return {}

    # Strip account_number from query — each account queries itself
    q_no_acc = HistoryQuery(
        from_date=q.from_date,
        to_date=q.to_date,
        min_amount=q.min_amount,
        max_amount=q.max_amount,
        direction=q.direction,
        counterparty=q.counterparty,
        description_search=q.description_search,
        category=q.category,
    )

    results: dict[str, list[Transaction]] = {}

    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_to_nick = {}
        for nick in nicknames:
            client = manager.get_client(nick)
            history = TransactionHistory(client)
            future = executor.submit(history.query, q_no_acc)
            future_to_nick[future] = nick

        for future in as_completed(future_to_nick):
            nick = future_to_nick[future]
            try:
                results[nick] = future.result()
            except Exception as exc:
                logger.warning("Failed to fetch history for %s: %s", nick, exc)
                results[nick] = []

    return results


# ──────────────────────────────────────────────────────────────────────────────
# Export functions
# ──────────────────────────────────────────────────────────────────────────────


_CSV_COLUMNS = [
    "timestamp",
    "account_number",
    "direction",
    "amount",
    "currency",
    "counterparty_name",
    "counterparty_account",
    "counterparty_bank",
    "description",
    "category",
]


def _tx_to_row(tx: Union[Transaction, CategorizedTransaction]) -> dict[str, Any]:
    """Convert a Transaction to a CSV-friendly dict."""
    row = {
        "timestamp": tx.timestamp.isoformat(),
        "account_number": tx.account_number,
        "direction": tx.type.value,
        "amount": str(tx.amount),
        "currency": tx.currency,
        "counterparty_name": tx.counterparty_name,
        "counterparty_account": tx.counterparty_account,
        "counterparty_bank": tx.counterparty_bank or "",
        "description": tx.description,
    }
    if isinstance(tx, CategorizedTransaction):
        row["category"] = tx.category
    else:
        row["category"] = ""
    return row


def export_to_csv(transactions: list[Transaction], output_path: Union[str, Path]) -> None:
    """Export transactions to a CSV file.

    Columns: timestamp, account_number, direction, amount, currency,
    counterparty_name, counterparty_account, counterparty_bank, description, category.

    Args:
        transactions: List to export.
        output_path: Destination .csv path.
    """
    output_path = Path(output_path)
    with output_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=_CSV_COLUMNS)
        writer.writeheader()
        for tx in transactions:
            writer.writerow(_tx_to_row(tx))


def export_to_json(transactions: list[Transaction], output_path: Union[str, Path]) -> None:
    """Export transactions to a JSON file (array of transaction dicts).

    Args:
        transactions: List to export.
        output_path: Destination .json path.
    """
    output_path = Path(output_path)
    data = [tx.to_dict() for tx in transactions]
    output_path.write_text(json.dumps(data, indent=2, ensure_ascii=False), encoding="utf-8")


def export_to_excel(
    transactions: list[Transaction],
    output_path: Union[str, Path],
    *,
    account_label: Optional[str] = None,
) -> None:
    """Export transactions to a formatted .xlsx file (requires ``openpyxl``).

    Columns: Date, Account, Direction, Amount, Currency, Counterparty Name,
    Counterparty Account, Counterparty Bank, Description, Category.

    Args:
        transactions: List to export.
        output_path: Destination .xlsx path.
        account_label: Optional account name shown in the header row.

    Raises:
        ImportError: If ``openpyxl`` is not installed. Install with
            ``pip install openpyxl``.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError as exc:
        raise ImportError(
            "export_to_excel requires the 'openpyxl' package. "
            "Install it with: pip install openpyxl"
        ) from exc

    output_path = Path(output_path)
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    # Header row
    headers = [
        "Date",
        "Account",
        "Direction",
        "Amount",
        "Currency",
        "Counterparty Name",
        "Counterparty Account",
        "Counterparty Bank",
        "Description",
        "Category",
    ]
    ws.append(headers)

    # Style header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # Data rows
    for tx in transactions:
        category = tx.category if isinstance(tx, CategorizedTransaction) else ""
        ws.append(
            [
                (
                    tx.timestamp.replace(tzinfo=None)
                    if hasattr(tx.timestamp, "tzinfo")
                    else tx.timestamp
                ),
                tx.account_number,
                tx.type.value,
                float(tx.amount),
                tx.currency,
                tx.counterparty_name,
                tx.counterparty_account,
                tx.counterparty_bank or "",
                tx.description,
                category,
            ]
        )

    # Auto-size columns (approximate)
    for col_idx, header in enumerate(headers, 1):
        max_len = len(header)
        for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
            for cell in row:
                if cell.value is not None:
                    max_len = max(max_len, min(len(str(cell.value)), 50))
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max_len + 2

    # Optional account label in row 1 col L (next to the table)
    if account_label:
        ws.cell(row=1, column=len(headers) + 2, value=f"Account: {account_label}").font = Font(
            italic=True, bold=True
        )

    wb.save(output_path)


__all__ = [
    "HistoryQuery",
    "CategorizedTransaction",
    "TransactionHistory",
    "query_all_accounts",
    "export_to_csv",
    "export_to_json",
    "export_to_excel",
]
